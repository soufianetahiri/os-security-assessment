import requests
import pandas as pd
from datetime import datetime, timedelta
import time
import logging
import sys
from tqdm import tqdm
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
import json
import os
import hashlib
from pathlib import Path
from typing import Dict, List, Optional, Any
from dataclasses import dataclass
import configparser

@dataclass
class RiskScoreWeights:
    cvss_score: float
    high_severity: float
    kev_status: float
    patch_availability: float
    eol_status: float
    latest_patch: float

@dataclass
class RiskAssessmentResult:
    risk_score: float
    high_severity_count: int
    high_severity_cves: List[str]
    kev_count: int
    kev_cves: List[str]
    avg_cvss: float

class OSRiskAssessment:
    def __init__(self, config_file: str = 'config.ini'):
        """
        Initialize the OS Risk Assessment tool.
        
        Args:
            config_file: Path to the configuration file
        """
        self.config = configparser.ConfigParser()
        self.config.read(config_file, encoding="utf-8")

        self.api_key = self.config.get('API', 'api_key')
        self.nvd_base_url = self.config.get('API', 'nvd_base_url')
        self.kev_url = self.config.get('API', 'kev_url')
        self.headers = {"apiKey": self.api_key}

        self.cache_dir = Path(self.config.get('Cache', 'cache_dir'))
        self.cache_duration = timedelta(days=int(self.config.get('Cache', 'cache_duration_days')))
        self.force_refresh = self.config.getboolean('Cache', 'force_refresh')

        self.weights = RiskScoreWeights(
            cvss_score=float(self.config.get('RiskWeights', 'cvss_score')),
            high_severity=float(self.config.get('RiskWeights', 'high_severity')),
            kev_status=float(self.config.get('RiskWeights', 'kev_status')),
            patch_availability=float(self.config.get('RiskWeights', 'patch_availability')),
            eol_status=float(self.config.get('RiskWeights', 'eol_status')),
            latest_patch=float(self.config.get('RiskWeights', 'latest_patch'))
        )

        self._setup_logging()
        self._init_cache()
        
        self.logger.info("Initializing OS Risk Assessment...")
        self.kev_catalog = self._get_kev_catalog()
        self.logger.info(f"Retrieved {len(self.kev_catalog)} KEV entries")
        
        self._load_static_data()
        self._verify_data_consistency()

    def _setup_logging(self) -> None:
        formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
        file_handler = logging.FileHandler('os_risk_assessment.log')
        file_handler.setFormatter(formatter)
        console_handler = logging.StreamHandler()
        console_handler.setFormatter(formatter)
        self.logger = logging.getLogger('OSRiskAssessment')
        self.logger.setLevel(logging.INFO)
        self.logger.handlers = []
        self.logger.addHandler(file_handler)
        self.logger.addHandler(console_handler)

    def _load_static_data(self) -> None:
        self.eol_dates = dict({key.strip().lower(): value.strip() for key, value in self.config.items('EOLDates')})
        self.patch_availability = dict({key.strip().lower(): value.strip() for key, value in self.config.items('PatchAvailability')})

    def _verify_data_consistency(self) -> None:
        all_os = set(list(self.eol_dates.keys()) + list(self.patch_availability.keys()))
        missing_eol = [os_name for os_name in all_os if os_name not in self.eol_dates]
        missing_patch = [os_name for os_name in all_os if os_name not in self.patch_availability]
        
        if missing_eol or missing_patch:
            error_msg = []
            if missing_eol:
                error_msg.append(f"Missing EOL dates for: {', '.join(missing_eol)}")
            if missing_patch:
                error_msg.append(f"Missing patch availability for: {', '.join(missing_patch)}")
            
            full_error = "Data consistency check failed! " + " ".join(error_msg)
            self.logger.error(full_error)
            raise ValueError(full_error)
        
        self.logger.info("Data consistency check passed successfully")

    def _init_cache(self) -> None:
        try:
            (self.cache_dir / 'cves').mkdir(parents=True, exist_ok=True)
            (self.cache_dir / 'kev').mkdir(parents=True, exist_ok=True)
            self._clean_old_cache_files()
        except Exception as e:
            self.logger.error(f"Error initializing cache: {str(e)}")
            raise

    def _clean_old_cache_files(self) -> None:
        try:
            current_time = datetime.now()
            for cache_type in ['cves', 'kev']:
                cache_path = self.cache_dir / cache_type
                for file_path in cache_path.glob('*.json'):
                    if current_time - datetime.fromtimestamp(file_path.stat().st_mtime) > self.cache_duration:
                        file_path.unlink()
                        self.logger.info(f"Removed old cache file: {file_path}")
        except Exception as e:
            self.logger.warning(f"Error cleaning cache files: {str(e)}")

    def _get_cache_path(self, os_name: str) -> Path:
        safe_name = hashlib.md5(os_name.encode()).hexdigest()
        return self.cache_dir / 'cves' / f"{safe_name}.json"

    def _get_kev_cache_path(self) -> Path:
        return self.cache_dir / 'kev' / 'kev_catalog.json'

    def _is_cache_valid(self, cache_path: Path) -> bool:
        if self.force_refresh:
            return False
        
        if not cache_path.exists():
            return False
        
        try:
            with cache_path.open('r') as f:
                json.load(f)
            return datetime.now() - datetime.fromtimestamp(cache_path.stat().st_mtime) < self.cache_duration
        except (json.JSONDecodeError, OSError):
            if cache_path.exists():
                cache_path.unlink()
            return False

    def _load_from_cache(self, cache_path: Path) -> Optional[List[Dict]]:
        try:
            with cache_path.open('r') as f:
                cache_data = json.load(f)
                
            if isinstance(cache_data, dict) and 'cves' in cache_data:
                cves = cache_data['cves']
                self.logger.info(f"Loaded {len(cves)} CVEs from cache created at {cache_data.get('timestamp', 'unknown')}")
                return cves
            else:
                self.logger.warning("Invalid cache data format")
                return None
                
        except Exception as e:
            self.logger.error(f"Error loading cache: {str(e)}")
            return None

    def _save_to_cache(self, cache_path: Path, data: List[Dict]) -> None:
        try:
            cache_data = {
                'timestamp': datetime.now().isoformat(),
                'cve_count': len(data),
                'cves': data
            }
            
            temp_path = cache_path.with_suffix('.tmp')
            with temp_path.open('w') as f:
                json.dump(cache_data, f)
            temp_path.replace(cache_path)
            self.logger.info(f"Cached {len(data)} CVEs to {cache_path}")
        except Exception as e:
            self.logger.error(f"Error saving to cache: {str(e)}")
            if temp_path.exists():
                temp_path.unlink()
                
    def _get_kev_catalog(self) -> List[Dict]:
        cache_path = self._get_kev_cache_path()
        validated_kevs = []
        
        if self._is_cache_valid(cache_path):
            self.logger.info("Loading KEV catalog from cache...")
            cached_data = self._load_from_cache(cache_path)
            if cached_data and isinstance(cached_data, dict) and 'kevs' in cached_data:
                raw_kevs = cached_data['kevs']
                self.logger.info(f"Found {len(raw_kevs)} KEV entries in cache")
            else:
                self.logger.warning("Invalid KEV cache format")
                raw_kevs = []
        else:
            self.logger.info("Fetching fresh KEV catalog from CISA...")
            try:
                response = requests.get(self.kev_url)
                if response.status_code == 200:
                    raw_kevs = response.json().get('vulnerabilities', [])
                    self.logger.info(f"Retrieved {len(raw_kevs)} KEV entries from CISA")
                else:
                    self.logger.error(f"Failed to fetch KEV catalog: HTTP {response.status_code}")
                    raw_kevs = []
            except Exception as e:
                self.logger.error(f"Error fetching KEV catalog: {str(e)}")
                raw_kevs = []
        
        if raw_kevs:
            self._save_to_cache(cache_path, {'timestamp': datetime.now().isoformat(), 'kevs': raw_kevs})
        
        return raw_kevs

    def _get_cpe_string(self, os_name: str) -> Optional[str]:
        """
        Get the CPE string for an operating system from the configuration file.
        
        Args:
            os_name: Name of the operating system
            
        Returns:
            Optional[str]: CPE string or None if not found
        """
        try:
            return self.config.get('CPENames', os_name)
        except configparser.NoOptionError:
            self.logger.error(f"CPE string not found for {os_name} in configuration file")
            return None

    def _fetch_nvd_cves(self, os_name: str) -> List[Dict]:
        cpe = self._get_cpe_string(os_name)
        if not cpe:
            self.logger.error(f"Could not create CPE string for {os_name}")
            return []

        self.logger.info(f"Searching CVEs for {os_name} using CPE: {cpe}")
        all_cves = []
        start_index = 0
        total_results = None
        pbar = None

        try:
            while total_results is None or start_index < total_results:
                params = {
                    'virtualMatchString': cpe,
                    'resultsPerPage': 50,
                    'startIndex': start_index
                }

                response = requests.get(
                    self.nvd_base_url,
                    headers=self.headers,
                    params=params
                )

                if response.status_code in (429, 503):
                    self.logger.warning("Rate limit reached, waiting 30 seconds...")
                    time.sleep(30)
                    continue

                if response.status_code != 200:
                    self.logger.error(f"Error fetching CVEs: {response.status_code}")
                    break

                data = response.json()

                if total_results is None:
                    total_results = data.get('totalResults', 0)
                    self.logger.info(f"Total CVEs found for {os_name}: {total_results}")
                    if total_results > 0:
                        pbar = tqdm(total=total_results, desc=f"Fetching CVEs for {os_name}", unit="CVE")

                vulnerabilities = data.get('vulnerabilities', [])
                current_batch = len(vulnerabilities)
                all_cves.extend(vulnerabilities)

                if pbar:
                    pbar.update(current_batch)
                    progress_percentage = (len(all_cves) / total_results) * 100
                    self.logger.info(f"Progress for {os_name}: {len(all_cves)}/{total_results} CVEs retrieved ({progress_percentage:.1f}%)")

                start_index += current_batch

                if current_batch < 50:
                    if pbar:
                        pbar.close()
                    break

                time.sleep(0.6)

            self.logger.info(f"CVE retrieval complete for {os_name}. Total CVEs retrieved: {len(all_cves)}")
            return all_cves

        except Exception as e:
            self.logger.error(f"Error retrieving CVEs for {os_name}: {str(e)}")
            if pbar:
                pbar.close()
            return []
    
    def validate_cve_for_os(self, cve: Dict, os_name: str) -> bool:
        """
        Validate if a CVE applies to the specified operating system using dynamic rules from the config file.
        
        Args:
            cve: CVE data dictionary
            os_name: Name of the operating system
            
        Returns:
            bool: True if CVE applies to the OS, False otherwise
        """
        try:
            configurations = cve.get('cve', {}).get('configurations', [])
            if not configurations:
                return True  # If no configuration data, include it
            
            # Get validation keywords for the OS from the config file
            validation_keywords = self.config.get('ValidationRules', os_name, fallback="").lower().split(",")
            if not validation_keywords:
                self.logger.warning(f"No validation keywords found for {os_name} in config file")
                return True  # Include CVE if no validation rules are defined
            
            # Check configurations against keywords
            for config in configurations:
                for node in config.get('nodes', []):
                    for match in node.get('cpeMatch', []):
                        cpe = match.get('criteria', '').lower()
                        if any(keyword in cpe for keyword in validation_keywords):
                            return True
            
            return False
        except Exception as e:
            self.logger.warning(f"Error validating CVE: {str(e)}")
            return True  # Include CVE if validation fails

    def _validate_kev_for_os(self, kev_entry: Dict, os_name: str) -> bool:
        """
        Validate if a KEV entry applies to the specified operating system using dynamic rules from the config file.
        
        Args:
            kev_entry: KEV vulnerability entry
            os_name: Name of the operating system
            
        Returns:
            bool: True if KEV entry applies to the OS, False otherwise
        """
        try:
            product = kev_entry.get('product', '').lower()
            vendor = kev_entry.get('vendorProject', '').lower()
            shortDescription = kev_entry.get('shortDescription', '').lower()
            
            # Get validation keywords for the OS from the config file
            validation_keywords = self.config.get('ValidationRules', os_name, fallback="").lower().split(",")
            if not validation_keywords:
                self.logger.warning(f"No validation keywords found for {os_name} in config file")
                return False  # Exclude KEV if no validation rules are defined
            
            # Check if any validation keyword matches the product or vendor
            return any(keyword in product or keyword in vendor or keyword in shortDescription for keyword in validation_keywords)
        except Exception as e:
            self.logger.warning(f"Error validating KEV entry: {str(e)}")
            return False  # Exclude KEV if validation fails

    def get_cves_for_os(self, os_name: str) -> List[Dict]:
        cache_path = self._get_cache_path(os_name)
        validated_cves = []
        
        if self._is_cache_valid(cache_path):
            self.logger.info(f"Loading CVEs for {os_name} from cache...")
            cached_data = self._load_from_cache(cache_path)
            if cached_data:
                self.logger.info(f"Found {len(cached_data)} CVEs in cache, validating...")
                raw_cves = cached_data
            else:
                self.logger.warning("Cache file exists but couldn't be loaded")
                raw_cves = []
        else:
            self.logger.info(f"Fetching fresh CVE data for {os_name}...")
            raw_cves = self._fetch_nvd_cves(os_name)
            
        total_cves = len(raw_cves)
        self.logger.info(f"Processing and validating {total_cves} CVEs...")
        
        with tqdm(total=total_cves, desc=f"Validating CVEs for {os_name}", unit="CVE") as pbar:
            for cve in raw_cves:
                try:
                    if 'cve' not in cve:
                        self.logger.debug("Skipping malformed CVE data: missing 'cve' key")
                        continue
                        
                    if self.validate_cve_for_os(cve, os_name):
                        validated_cves.append(cve)
                    else:
                        self.logger.debug(f"Skipping CVE {cve['cve'].get('id', 'unknown')} - not applicable to {os_name}")
                except Exception as e:
                    self.logger.warning(f"Error processing CVE: {str(e)}")
                finally:
                    pbar.update(1)

        if not self._is_cache_valid(cache_path) and validated_cves:
            self._save_to_cache(cache_path, validated_cves)
            
        validation_ratio = (len(validated_cves) / total_cves * 100) if total_cves > 0 else 0
        self.logger.info(
            f"CVE validation complete for {os_name}:\n"
            f"- Total CVEs processed: {total_cves}\n"
            f"- Applicable CVEs: {len(validated_cves)} ({validation_ratio:.1f}%)"
        )
        
        return validated_cves

    def _check_kev_status(self, cve_id: str, os_name: str) -> bool:
        try:
            for kev in self.kev_catalog:
                if kev['cveID'] == cve_id:
                    if self._validate_kev_for_os(kev, os_name):
                        self.logger.info(f"KEV found and validated {kev}")
                        return True
            return False
        except Exception as e:
            self.logger.warning(f"Error checking KEV status for {cve_id}: {str(e)}")
            return False

    def calculate_risk_score(self, os_name: str, cves: List[Dict]) -> RiskAssessmentResult:
        self.logger.info(f"Calculating risk score for {os_name}...")
        self.logger.info(f"Available EOL dates: {list(self.eol_dates.values())}")

        try:
            cvss_scores = []
            high_severity_count = 0
            high_severity_cves = []
            kev_count = 0
            kev_cves = []

            for cve in tqdm(cves, desc=f"Analyzing CVEs for {os_name}", unit="CVE"):
                cve_id = cve['cve']['id']
                metrics = cve.get('cve', {}).get('metrics', {}).get('cvssMetricV31', [])
                
                if metrics:
                    base_score = metrics[0].get('cvssData', {}).get('baseScore', 0)
                    cvss_scores.append(base_score)
                    
                    if base_score >= 7.0:
                        high_severity_count += 1
                        high_severity_cves.append(cve_id)
                    
                    if self._check_kev_status(cve_id, os_name):
                        kev_count += 1
                        kev_cves.append(cve_id)

            avg_cvss = sum(cvss_scores) / len(cvss_scores) if cvss_scores else 0
            cvss_component = (avg_cvss / 10.0) * self.weights.cvss_score * 100
            high_severity_component = min(high_severity_count / 10.0, 1.0) * self.weights.high_severity * 100
            kev_component = min(kev_count / 5.0, 1.0) * self.weights.kev_status * 100

            patch_scores = {'Regular': 0.2, 'Limited': 0.6, 'None': 1.0}
            patch_status = self.patch_availability.get(os_name, 'None')
            patch_component = patch_scores[patch_status] * self.weights.patch_availability * 100

            eol_date = datetime.strptime(self.eol_dates[os_name.lower()], "%Y-%m-%d")
            days_until_eol = (eol_date - datetime.now()).days
            eol_component = max(0, min(1.0, 1 - (days_until_eol / 365))) * self.weights.eol_status * 100

            latest_patch_component = 0 if patch_status == 'Regular' else self.weights.latest_patch * 100

            total_risk = min(100, cvss_component + high_severity_component + kev_component + 
                            patch_component + eol_component + latest_patch_component)

            self.logger.info(f"Risk calculation complete for {os_name} - Score: {total_risk:.2f}")

            return RiskAssessmentResult(
                risk_score=total_risk,
                high_severity_count=high_severity_count,
                high_severity_cves=high_severity_cves,
                kev_count=kev_count,
                kev_cves=kev_cves,
                avg_cvss=avg_cvss
            )

        except Exception as e:
            self.logger.error(f"Error calculating risk score for {os_name}: {str(e)}")
            raise

    def generate_report(self, operating_systems: List[str]) -> List[Dict]:
        self.logger.info("Starting report generation...")
        results = []

        try:
            for os_name in tqdm(operating_systems, desc="Processing Operating Systems", unit="OS"):
                self.logger.info(f"\nProcessing {os_name}...")
                cves = self.get_cves_for_os(os_name)
                risk_data = self.calculate_risk_score(os_name.lower(), cves)

                results.append({
                    'Operating System': os_name.lower(),
                    'Known CVEs': len(cves),
                    'High Severity CVEs': risk_data.high_severity_count,
                    'KEV Count': risk_data.kev_count,
                    'Patch Availability': self.patch_availability[os_name.lower()],
                    'EoL/EoSupport': self.eol_dates[os_name.lower()],
                    'Security Risk Score': round(risk_data.risk_score, 2),
                    'High Severity CVEs List': ', '.join(risk_data.high_severity_cves),
                    'KEV List': ', '.join(risk_data.kev_cves)
                })

            self._create_excel_report(results)
            self.logger.info("Report generation complete!")
            return results

        except Exception as e:
            self.logger.error(f"Error generating report: {str(e)}")
            raise

    def _create_excel_report(self, results: List[Dict]) -> None:
        try:
            df = pd.DataFrame(results)
            writer = pd.ExcelWriter('os_security_assessment.xlsx', engine='openpyxl')
            df.to_excel(writer, index=False, sheet_name='Risk Assessment')

            workbook = writer.book
            worksheet = writer.sheets['Risk Assessment']

            for idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=len(results)+1, min_col=7, max_col=7)):
                cell = row[0]
                score = df['Security Risk Score'].iloc[idx]
                
                if score >= 80:
                    cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                elif score >= 60:
                    cell.fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
                elif score >= 40:
                    cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                else:
                    cell.fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')

            for column in worksheet.columns:
                max_length = 0
                column = list(column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

            writer.close()

        except Exception as e:
            self.logger.error(f"Error creating Excel report: {str(e)}")
            raise

def main():
    try:
        assessment = OSRiskAssessment(config_file='config.ini')
        operating_systems = [os.strip() for os in assessment.config.get('OperatingSystems', 'operating_systems').split(',')]
        results = assessment.generate_report(operating_systems)
        print("Assessment complete. Results saved to 'os_security_assessment.xlsx'")
    
    except Exception as e:
        print(f"Error during assessment: {str(e)}")
        logging.error(f"Error during assessment: {str(e)}", exc_info=True)
        sys.exit(1)

if __name__ == "__main__":
    main()
